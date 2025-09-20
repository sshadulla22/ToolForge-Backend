from fastapi import FastAPI, UploadFile, File, Form
from fastapi.responses import FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from pdf2docx import Converter
from PyPDF2 import PdfReader, PdfWriter
from PIL import Image, ImageDraw, ImageFont
import io, os, zipfile, json, base64, qrcode
from pydantic import BaseModel
from fastapi.responses import StreamingResponse
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import FileResponse
import tempfile, os, subprocess, shutil
from uuid import uuid4
from pdf2docx import Converter
from pdf2image import convert_from_path
from PIL import Image
from docx2pdf import convert as docx2pdf_convert
import zipfile
import aspose.slides as slides
from docx import Document
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
import tempfile
import io
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from fastapi.responses import StreamingResponse
import fitz  # PyMuPDF
from fastapi.responses import RedirectResponse


# ---------------- App Setup ----------------
app = FastAPI(title="ToolForge Backend API 🚀")

UPLOAD_DIR = "temp"
os.makedirs(UPLOAD_DIR, exist_ok=True)

# ---------------- CORS ----------------
app.add_middleware(
    CORSMiddleware,
    allow_origins=["https://toolforge-frontend.onrender.com"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------------- Utility ----------------
async def save_upload(file: UploadFile) -> str:
    """Save uploaded file to temp directory with unique name"""
    ext = os.path.splitext(file.filename)[1]
    unique_name = f"{uuid4().hex}{ext}"
    file_path = os.path.join(UPLOAD_DIR, unique_name)
    content = await file.read()
    with open(file_path, "wb") as f:
        f.write(content)
    return file_path

# ---------------- Root ----------------
@app.get("/", include_in_schema=False)
async def root():
    return {
        "message": "Welcome to ToolForge Backend 🚀",
        "docs": "/docs",
        "redoc": "/redoc"
    }



# ---------------- PDF → DOCX ----------------
@app.post("/pdf-to-docx/")
async def pdf_to_docx(file: UploadFile = File(...)):
    file_path = f"temp/{file.filename}"
    with open(file_path, "wb") as f:
        f.write(await file.read())
    output_file = file_path.replace(".pdf", ".docx")
    cv = Converter(file_path)
    cv.convert(output_file, start=0, end=None)
    cv.close()
    return FileResponse(output_file, filename="converted.docx")

# Merge PDFs
@app.post("/merge-pdf/")
async def merge_pdf(files: list[UploadFile] = File(...)):
    merger = PdfWriter()
    for file in files:
        reader = PdfReader(io.BytesIO(await file.read()))
        for page in reader.pages:
            merger.add_page(page)
    output_file = "temp/merged.pdf"
    merger.write(output_file)
    merger.close()
    return FileResponse(output_file, media_type="application/pdf", filename="merged.pdf")

@app.post("/split-pdf/")
async def split_pdf(file: UploadFile = File(...), pages_per_split: int = Form(...)):
    reader = PdfReader(io.BytesIO(await file.read()))
    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, "w") as zip_file:
        for i in range(0, len(reader.pages), pages_per_split):
            writer = PdfWriter()
            for j in range(i, min(i+pages_per_split, len(reader.pages))):
                writer.add_page(reader.pages[j])
            buf = io.BytesIO()
            writer.write(buf)
            buf.seek(0)
            zip_file.writestr(f"split_{i+1}.pdf", buf.read())

    zip_buffer.seek(0)
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=split_pdfs.zip"}
    )

# Extract Text from PDF
@app.post("/extract-text/")
async def extract_text(file: UploadFile = File(...)):
    reader = PdfReader(io.BytesIO(await file.read()))
    text = "".join([page.extract_text() or "" for page in reader.pages])
    return JSONResponse({"text": text})

# ---------------- Resize Image ----------------
@app.post("/resize-image/")
async def resize_image(file: UploadFile = File(...), width: int = Form(...), height: int = Form(...)):
    image = Image.open(io.BytesIO(await file.read()))
    resized = image.resize((width, height))
    buf = io.BytesIO()
    resized.save(buf, format="JPEG")
    buf.seek(0)
    return StreamingResponse(buf, media_type="image/jpeg", headers={"Content-Disposition": "attachment; filename=resized.jpg"})

# ---------------- Convert Format ----------------
@app.post("/convert-format/")
async def convert_format(file: UploadFile = File(...), format: str = Form(...)):
    image = Image.open(io.BytesIO(await file.read()))
    buf = io.BytesIO()
    image.save(buf, format=format)
    buf.seek(0)
    return StreamingResponse(buf, media_type=f"image/{format.lower()}", headers={"Content-Disposition": f"attachment; filename=converted.{format.lower()}"})

# ---------------- Add Watermark ----------------
@app.post("/watermark/")
async def add_watermark(file: UploadFile = File(...), text: str = Form(...), opacity: int = Form(...), font_size: int = Form(...)):
    image = Image.open(io.BytesIO(await file.read()))
    if image.mode in ("RGBA","LA"):
        image = image.convert("RGB")
    watermark = Image.new("RGBA", image.size, (0,0,0,0))
    draw = ImageDraw.Draw(watermark)
    try:
        font = ImageFont.truetype("arial.ttf", font_size)
    except:
        font = ImageFont.load_default()
    step_x = font_size * 10
    step_y = font_size * 8
    for y in range(0, image.height, step_y):
        for x in range(0, image.width, step_x):
            draw.text((x, y), text, fill=(255,255,255,opacity), font=font)
    watermarked = Image.alpha_composite(image.convert("RGBA"), watermark)
    buf = io.BytesIO()
    watermarked.convert("RGB").save(buf, format="JPEG")
    buf.seek(0)
    return StreamingResponse(buf, media_type="image/jpeg", headers={"Content-Disposition": "attachment; filename=watermarked.jpg"})

# ---------------- Compress Image ----------------
@app.post("/compress-image/")
async def compress_image(file: UploadFile = File(...), target_size: int = Form(...)):
    image = Image.open(io.BytesIO(await file.read()))
    if image.mode in ("RGBA","LA"):
        image = image.convert("RGB")
    quality = 95
    buf = io.BytesIO()
    image.save(buf, format="JPEG", quality=quality)
    compressed_data = buf.getvalue()
    compressed_size = len(compressed_data)/1024
    while compressed_size > target_size and quality > 10:
        quality -= 5
        buf = io.BytesIO()
        image.save(buf, format="JPEG", quality=quality)
        compressed_data = buf.getvalue()
        compressed_size = len(compressed_data)/1024
    buf.seek(0)
    return StreamingResponse(buf, media_type="image/jpeg", headers={"Content-Disposition": "attachment; filename=compressed.jpg"})

# ---------------- Generate QR ----------------
@app.post("/generate-qr/")
async def generate_qr(text: str = Form(...)):
    qr_img = qrcode.make(text)
    buf = io.BytesIO()
    qr_img.save(buf, format="PNG")
    buf.seek(0)
    return StreamingResponse(buf, media_type="image/png", headers={"Content-Disposition": "attachment; filename=qrcode.png"})

# ---------------- Split PDF ----------------
@app.post("/split-pdf/")
async def split_pdf(file: UploadFile = File(...), pages_per_split: int = Form(...)):
    reader = PdfReader(io.BytesIO(await file.read()))
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w") as zip_file:
        for i in range(0, len(reader.pages), pages_per_split):
            writer = PdfWriter()
            for j in range(i, min(i+pages_per_split, len(reader.pages))):
                writer.add_page(reader.pages[j])
            buf = io.BytesIO()
            writer.write(buf)
            buf.seek(0)
            zip_file.writestr(f"split_{i+1}.pdf", buf.read())
    zip_buffer.seek(0)
    return StreamingResponse(zip_buffer, media_type="application/zip", headers={"Content-Disposition": "attachment; filename=split_pdfs.zip"})

# ---------------- Base64 Decode ----------------
@app.post("/base64-decode/")
async def base64_decode(encoded: str = Form(...)):
    data = base64.b64decode(encoded)
    buf = io.BytesIO(data)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/octet-stream", headers={"Content-Disposition": "attachment; filename=decoded.bin"})
# QR Code Generator
@app.post("/generate-qr/")
async def generate_qr(text: str = Form(...)):
    qr_img = qrcode.make(text)
    buf = io.BytesIO()
    qr_img.save(buf, format="PNG")
    buf.seek(0)
    return FileResponse(buf, media_type="image/png", filename="qrcode.png")

# Base64 Encode
@app.post("/base64-encode/")
async def base64_encode(file: UploadFile = File(None), text: str = Form(None)):
    if not file and not text:
        return {"error": "No input provided"}
    data = await file.read() if file else text.encode()
    return {"base64": base64.b64encode(data).decode()}

# Base64 Decode
@app.post("/base64-decode/")
async def base64_decode(encoded: str = Form(...)):
    data = base64.b64decode(encoded)
    buf = io.BytesIO(data)
    buf.seek(0)
    return FileResponse(buf, media_type="application/octet-stream", filename="decoded.bin")

# Pydantic model for JSON input
class JsonInput(BaseModel):
    json_text: str

# JSON Formatter endpoint
@app.post("/format-json/")
async def format_json(data: JsonInput):
    try:
        parsed = json.loads(data.json_text)
        return {"formatted": json.dumps(parsed, indent=4)}
    except json.JSONDecodeError:
        return {"error": "Invalid JSON"}



# ---------- Conversions ---------- #

# ---------------- PDF → DOCX ----------------
@app.post("/convert/pdf-to-docx")
async def convert_pdf_to_docx(file: UploadFile = File(...)):
    input_path = await save_upload(file)
    output_path = input_path.replace(".pdf", ".docx")
    try:
        cv = Converter(input_path)
        cv.convert(output_path, start=0, end=None)
        cv.close()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    return FileResponse(output_path, filename=os.path.basename(output_path))

# ---------------- DOCX → PDF ----------------
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import FileResponse
from docx2pdf import convert as docx2pdf_convert
import os

@app.post("/convert/docx-to-pdf")
async def convert_docx_to_pdf(file: UploadFile = File(...)):
    input_path = os.path.join(UPLOAD_DIR, file.filename)
    with open(input_path, "wb") as f:
        f.write(await file.read())

    output_path = os.path.splitext(input_path)[0] + ".pdf"

    try:
        doc = Document(input_path)
        c = canvas.Canvas(output_path, pagesize=A4)
        width, height = A4
        y = height - 50  # Start from top

        for para in doc.paragraphs:
            text = para.text
            if y < 50:  # Start new page
                c.showPage()
                y = height - 50
            c.drawString(50, y, text)
            y -= 15  # line spacing

        c.save()

        if not os.path.exists(output_path):
            raise Exception("Conversion failed: PDF not created")

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"DOCX → PDF conversion failed: {e}")

    return FileResponse(output_path, filename=os.path.basename(output_path))


# ....................PFG_IMAGE...............
@app.post("/convert/pdf-to-image")
async def convert_pdf_to_image(file: UploadFile = File(...), format: str = Form("jpg")):
    try:
        # Save uploaded file
        temp_dir = tempfile.mkdtemp()
        input_path = os.path.join(temp_dir, file.filename)
        with open(input_path, "wb") as f:
            f.write(await file.read())

        # Convert PDF to images
        images = convert_from_path(input_path)
        output_files = []

        format_mapping = {"jpg": "JPEG", "png": "PNG"}
        for i, img in enumerate(images):
            img = img.convert("RGB")  # important for JPEG
            out_path = os.path.join(temp_dir, f"page_{i}.{format.lower()}")
            img.save(out_path, format_mapping.get(format.lower(), "JPEG"))
            output_files.append(out_path)

        # Single image -> return file, multiple -> zip
        if len(output_files) == 1:
            return FileResponse(output_files[0], media_type=f"image/{format.lower()}", filename=f"page_0.{format.lower()}")
        else:
            zip_path = os.path.join(temp_dir, "images.zip")
            with zipfile.ZipFile(zip_path, "w") as zipf:
                for f in output_files:
                    zipf.write(f, os.path.basename(f))
            return FileResponse(zip_path, media_type="application/zip", filename="images.zip")

    except Exception as e:
        return {"error": "Conversion failed. Please try again.", "details": str(e)}

# ---------------- Image → PDF ----------------
@app.post("/convert/image-to-pdf")
async def convert_image_to_pdf(file: UploadFile = File(...)):
    input_path = await save_upload(file)
    output_path = input_path + ".pdf"
    try:
        img = Image.open(input_path)
        img.convert("RGB").save(output_path)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    return FileResponse(output_path, filename=os.path.basename(output_path))

# ---------------- PPT/PPTX → PDF ----------------
@app.post("/convert/ppt-to-pdf")
async def convert_ppt_to_pdf(file: UploadFile = File(...)):
    input_path = os.path.join(UPLOAD_DIR, file.filename)
    with open(input_path, "wb") as f:
        f.write(await file.read())

    try:
        # Load presentation
        pres = slides.Presentation(input_path)

        # Save to PDF
        output_path = os.path.splitext(input_path)[0] + ".pdf"
        pres.save(output_path, slides.export.SaveFormat.PDF)

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PPT/PPTX → PDF conversion failed: {e}")

    return FileResponse(output_path, filename=os.path.basename(output_path))

@app.post("/convert/excel-to-pdf")
async def convert_excel_to_pdf(file: UploadFile = File(...)):
    input_path = os.path.join(UPLOAD_DIR, file.filename)
    with open(input_path, "wb") as f:
        f.write(await file.read())

    output_path = os.path.splitext(input_path)[0] + ".pdf"

    try:
        wb = load_workbook(input_path)
        c = canvas.Canvas(output_path, pagesize=A4)
        width, height = A4
        margin = 50
        y = height - margin

        for sheet in wb.worksheets:
            c.setFont("Helvetica-Bold", 14)
            c.drawString(margin, y, f"Sheet: {sheet.title}")
            y -= 20
            c.setFont("Helvetica", 12)

            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                c.drawString(margin, y, row_text)
                y -= 15
                if y < margin:
                    c.showPage()
                    y = height - margin

            c.showPage()  # New page for next sheet

        c.save()

        if not os.path.exists(output_path):
            raise Exception("PDF not created")

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel → PDF conversion failed: {e}")

    return FileResponse(output_path, filename=os.path.basename(output_path))

@app.post("/compress-pdf/")
async def compress_pdf(file: UploadFile = File(...), level: str = Form("medium")):
    try:
        pdf_bytes = await file.read()
        pdf_doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        compressed_pdf = io.BytesIO()

        # Map compression levels to image quality (0-100)
        quality_map = {
            "high": 20,
            "medium": 50,
            "low": 80
        }
        quality = quality_map.get(level.lower(), 50)

        # Save compressed PDF using PyMuPDF
        pdf_doc.save(
            compressed_pdf,
            garbage=4,       # remove unused objects
            deflate=True,    # compress streams
            clean=True       # clean up
        )
        compressed_pdf.seek(0)

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PDF compression failed: {e}")

    return StreamingResponse(
        compressed_pdf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=compressed_{file.filename}"}
    )
